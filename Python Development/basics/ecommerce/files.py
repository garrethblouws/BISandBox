import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
sheet.cell(1,1)
print(sheet.max_row)
